"""
main.py - FastAPI server for CV Matcher system.
Run with: uvicorn main:app --reload --port 8000
"""

import asyncio
import hashlib
import json
import logging
import os
import re
import smtplib
import sys
import threading
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Optional

from fastapi import FastAPI, HTTPException, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse, Response, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from jose import JWTError, jwt
from passlib.context import CryptContext
from datetime import timedelta

import time

import database
import cv_scanner
import cv_processor
import job_matcher
import quick_matcher
import civi_scraper


def _claude_with_retry(fn, max_retries: int = 4):
    """
    Call fn() (which must invoke claude.messages.create) and retry automatically
    when Anthropic returns a 529 (overloaded) or 429 (rate-limit) status.
    Waits 3 s → 6 s → 12 s between attempts (exponential back-off).
    Raises HTTPException(503) with a Hebrew message if all retries fail.
    """
    from anthropic import APIStatusError
    for attempt in range(max_retries):
        try:
            return fn()
        except APIStatusError as exc:
            if exc.status_code not in (429, 529):
                raise   # not a retryable error — propagate immediately
            if attempt < max_retries - 1:
                wait = 3 * (2 ** attempt)   # 3 s → 6 s → 12 s
                logging.warning("Anthropic %s (attempt %d/%d) — retrying in %ss",
                                exc.status_code, attempt + 1, max_retries, wait)
                time.sleep(wait)
            else:
                logging.error("Anthropic %s — all %d retries exhausted", exc.status_code, max_retries)
                raise HTTPException(
                    status_code=503,
                    detail="שרת ה-AI עמוס כרגע. אנא המתן מספר שניות ונסה שוב.",
                )

LOG_PATH = os.path.join(os.path.dirname(__file__), "app.log")
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8", mode="w"),
        logging.StreamHandler(),
    ],
)
# Silence noisy third-party loggers
for _noisy in ("pdfminer", "pdfminer.pdffont", "pdfminer.pdfpage", "pdfminer.converter",
                "httpcore", "httpx", "anthropic._base_client"):
    logging.getLogger(_noisy).setLevel(logging.ERROR)
logger = logging.getLogger(__name__)

app = FastAPI(title="CV Matcher API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)

database.init_db()

# ─── Auth config ──────────────────────────────────────────────────────────────

_SECRET = os.environ.get("JWT_SECRET", "aih-secret-key-change-in-production")
_ALGO   = "HS256"
_DAYS   = 7
_pwd    = CryptContext(schemes=["bcrypt"], deprecated="auto")


def _make_token(user_id: int, email: str) -> str:
    exp = datetime.utcnow() + timedelta(days=_DAYS)
    return jwt.encode({"user_id": user_id, "email": email, "exp": exp}, _SECRET, algorithm=_ALGO)


def _decode_token(token: str) -> dict | None:
    try:
        return jwt.decode(token, _SECRET, algorithms=[_ALGO])
    except JWTError:
        return None


def _get_token(request: Request) -> str:
    token = request.cookies.get("aih_token", "")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    return token


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    # Skip non-API routes and OPTIONS preflight
    if not path.startswith("/api/") or request.method == "OPTIONS":
        return await call_next(request)
    # Public endpoints
    if path == "/api/auth/login":
        return await call_next(request)
    # If no user has a password yet — allow all (initial setup)
    if not database.any_user_has_password():
        return await call_next(request)
    # Verify token
    token = _get_token(request)
    if not token or not _decode_token(token):
        return JSONResponse({"detail": "לא מורשה"}, status_code=401)
    return await call_next(request)


# ─── Email helpers ───────────────────────────────────────────────────────────

def _send_reminder_email(to_email: str, to_name: str, reminder: dict) -> str | None:
    """Send a single reminder email. Returns None on success, or an error string on failure."""
    smtp_host = database.get_setting("smtp_host", "smtp.gmail.com")
    smtp_port = int(database.get_setting("smtp_port", "587"))
    smtp_user = database.get_setting("smtp_user", "")
    smtp_pass = database.get_setting("smtp_pass", "")
    from_name = database.get_setting("smtp_from_name", "AI Head Hunter")

    if not smtp_user or not smtp_pass:
        return "SMTP לא מוגדר — חסרים כתובת מייל או סיסמה"

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"🔔 תזכורת: {reminder['title']}"
        msg["From"] = f"{from_name} <{smtp_user}>"
        msg["To"] = f"{to_name} <{to_email}>"

        body_parts = [f"<h2>🔔 {reminder['title']}</h2>"]
        if reminder.get("description"):
            body_parts.append(f"<p>{reminder['description']}</p>")
        if reminder.get("candidate_name"):
            body_parts.append(f"<p>👤 <b>מועמד:</b> {reminder['candidate_name']}</p>")
        if reminder.get("job_title"):
            body_parts.append(f"<p>💼 <b>משרה:</b> {reminder['job_title']}</p>")
        body_parts.append(f"<p style='color:#888;font-size:12px'>נשלח על ידי AI Head Hunter</p>")
        html_body = f"<div dir='rtl' style='font-family:sans-serif'>{''.join(body_parts)}</div>"

        msg.attach(MIMEText(html_body, "html", "utf-8"))

        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, [to_email], msg.as_string())

        logger.info("Reminder email sent to %s for reminder %s", to_email, reminder["id"])
        return None  # success
    except Exception as exc:
        err = str(exc)
        logger.error("Failed to send reminder email to %s: %s", to_email, err)
        return err


def _email_checker_loop():
    """Background thread: check every 60s for due reminders with assigned users and send emails."""
    import time
    while True:
        try:
            due = database.get_due_reminders_for_email()
            for r in due:
                err = _send_reminder_email(r["user_email"], r["user_name"], r)
                if err is None:
                    database.mark_email_sent(r["id"])
                else:
                    database.mark_email_failed(r["id"], err)
        except Exception as exc:
            logger.error("Email checker error: %s", exc)
        time.sleep(60)


# Start background email thread
_email_thread = threading.Thread(target=_email_checker_loop, daemon=True)
_email_thread.start()


# ─── Request / Response models ──────────────────────────────────────────────

class ScanRequest(BaseModel):
    folder_path: str


class JobRequirements(BaseModel):
    role_title: str
    min_experience_years: Optional[float] = 0
    required_technologies: Optional[list[str]] = []
    nice_to_have: Optional[list[str]] = []
    management_required: Optional[bool] = False
    location: Optional[str] = None
    domain: Optional[str] = None
    org_type: Optional[str] = None           # לקוח קצה / אינטגרטור / ספק / פיננסי / ממשלתי / סטארטאפ / אנטרפרייז
    salary_range: Optional[str] = None       # e.g. "25-35K + רכב"
    hybrid_mode: Optional[str] = None        # משרדי / היברידי / מרחוק
    vendor_experience_required: Optional[bool] = False
    additional_notes: Optional[str] = None
    top_n: Optional[int] = None
    min_match_pct: Optional[int] = None


class MatchRequest(BaseModel):
    job_requirements: JobRequirements
    top_n: Optional[int] = None        # None = use job's saved setting or default
    min_match_pct: Optional[int] = None  # None = use job's saved setting or default
    candidate_ids: Optional[list[int]] = None  # if set, only match these candidates


# ─── Endpoints ──────────────────────────────────────────────────────────────

@app.get("/api/browse-folder")
def browse_folder():
    """Open a native Windows folder picker dialog and return the selected path."""
    import subprocess
    ps_script = (
        "Add-Type -AssemblyName System.Windows.Forms; "
        "$d = New-Object System.Windows.Forms.FolderBrowserDialog; "
        "$d.Description = 'בחר תיקיית קורות חיים'; "
        "$d.RootFolder = 'MyComputer'; "
        "$d.ShowNewFolderButton = $false; "
        "$null = $d.ShowDialog(); "
        "Write-Output $d.SelectedPath"
    )
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps_script],
            capture_output=True, text=True, timeout=120, encoding="utf-8"
        )
        path = result.stdout.strip()
        if path:
            return {"path": path}
        return {"path": None}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/stats")
def get_stats():
    return database.get_stats()


@app.get("/api/candidates")
def list_candidates():
    return database.get_all_candidates()


@app.post("/api/scan")
def scan_folder(req: ScanRequest):
    """
    Scan folder, extract and store new candidates.
    Returns a streaming Server-Sent Events response with progress updates.
    """
    folder = req.folder_path

    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Folder not found: {folder}")

    def generate():
        started_at = datetime.utcnow().isoformat()
        files_found = cv_scanner.count_files(folder)
        files_processed = 0
        duplicates_skipped = 0
        already_known = 0
        errors = 0

        # Load all known hashes once — delta mode: skip text extraction for known files
        known_hashes = database.get_all_hashes()

        yield f"data: {json.dumps({'type': 'start', 'files_found': files_found, 'already_in_db': len(known_hashes)})}\n\n"

        for item in cv_scanner.scan_folder(folder, known_hashes=known_hashes):
            file_name = item["file_name"]
            file_path = item["file_path"]
            md5 = item["md5"]
            text = item["text"]
            error = item["error"]

            if error:
                errors += 1
                yield f"data: {json.dumps({'type': 'error', 'file': file_name, 'error': error})}\n\n"
                continue

            # Already in DB — skip silently (no text was extracted, very fast)
            if item.get("already_known"):
                already_known += 1
                continue

            if not text.strip():
                errors += 1
                yield f"data: {json.dumps({'type': 'error', 'file': file_name, 'error': 'Could not extract text'})}\n\n"
                continue

            # Extract structured data via Claude
            yield f"data: {json.dumps({'type': 'processing', 'file': file_name})}\n\n"
            extracted = cv_processor.extract_candidate_data(text, file_name)

            if not extracted:
                errors += 1
                yield f"data: {json.dumps({'type': 'error', 'file': file_name, 'error': 'Extraction failed'})}\n\n"
                continue

            # Check near-duplicate (same name + email)
            name = extracted.get("name") or ""
            email = extracted.get("email") or ""
            existing = database.candidate_exists_by_identity(name, email)
            if existing:
                duplicates_skipped += 1
                dup_reason = f"Same candidate as {existing['file_name']}"
                yield f"data: {json.dumps({'type': 'duplicate', 'file': file_name, 'reason': dup_reason})}\n\n"
                # Still store hash so we don't re-process this file
                database.insert_candidate({
                    "file_hash": md5,
                    "file_path": file_path,
                    "file_name": file_name,
                    "name": name + " (dup)",  # mark as dup so it won't match identity check
                })
                continue

            extracted["file_hash"] = md5
            extracted["file_path"] = file_path
            extracted["file_name"] = file_name

            database.insert_candidate(extracted)
            files_processed += 1

            yield f"data: {json.dumps({'type': 'done', 'file': file_name, 'name': extracted.get('name', 'Unknown')})}\n\n"

        database.log_scan(
            folder, files_found, files_processed, duplicates_skipped, errors, started_at
        )

        yield f"data: {json.dumps({'type': 'finished', 'files_found': files_found, 'already_known': already_known, 'files_processed': files_processed, 'duplicates_skipped': duplicates_skipped, 'errors': errors})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


@app.post("/api/import-excel")
async def import_from_excel(file: UploadFile = File(...)):
    """
    ייבוא מועמדים מקובץ האקסל הקיים (שנוצר על ידי הסקריפט הקודם).
    פורמט עמודות: Name, Current_Role, Years_Experience, Technologies,
                  Managerial_Experience, Domain, Summary, Phone, Email, שם קובץ, לינק מהיר
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="חסרה ספריית openpyxl. הרץ: pip install openpyxl")

    content = await file.read()

    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    def col(row, name):
        try:
            idx = headers.index(name)
            v = row[idx].value
            return str(v).strip() if v and str(v) != "לא צוין" else None
        except (ValueError, IndexError):
            return None

    def get_file_path(row):
        """חלץ נתיב קובץ מעמודת לינק מהיר - תומך בנוסחת HYPERLINK ובהיפרקישורים של Excel."""
        try:
            idx = headers.index("לינק מהיר")
            cell = row[idx]
            # נוסחת HYPERLINK כטקסט: =HYPERLINK("G:\path","label")
            val = str(cell.value or "")
            m = re.search(r'HYPERLINK\("([^"]+)"', val, re.IGNORECASE)
            if m:
                return m.group(1)
            # היפרקישור של Excel (לא נוסחה)
            if cell.hyperlink and cell.hyperlink.target:
                t = cell.hyperlink.target
                if t.startswith(("G:\\", "C:\\", "D:\\", "G:/", "C:/", "D:/")):
                    return t.replace("/", "\\")
            # ערך גולמי שהוא נתיב
            if val.startswith(("G:\\", "C:\\", "D:\\")):
                return val
        except (ValueError, IndexError):
            pass
        # fallback: שם קובץ + נתיב ידוע
        file_name = col(row, "שם קובץ") or ""
        if file_name:
            return r"G:\האחסון שלי\connectech\גיוס\connectech civi\ConnecTech (9)\ConnecTech" + "\\" + file_name
        return ""

    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2):
        name = col(row, "Name")
        if not name:
            continue

        email = col(row, "Email")
        phone = col(row, "Phone")

        # dedup by name+email
        if email and database.candidate_exists_by_identity(name, email):
            skipped += 1
            continue

        # build a fake hash from name+email+phone so we don't double-import
        hash_src = f"{name}|{email}|{phone}".encode()
        fake_hash = hashlib.md5(hash_src).hexdigest()

        if database.hash_exists(fake_hash):
            skipped += 1
            continue

        years_raw = col(row, "Years_Experience")
        years = None
        if years_raw:
            m = re.search(r"(\d+(?:\.\d+)?)", years_raw)
            years = float(m.group(1)) if m else None

        techs_raw = col(row, "Technologies") or ""
        techs = [t.strip() for t in re.split(r"[,;\n|/]", techs_raw) if t.strip()]

        mgmt_raw = (col(row, "Managerial_Experience") or "").lower()
        mgmt = "כן" in mgmt_raw or "yes" in mgmt_raw or "true" in mgmt_raw

        # extract file path from HYPERLINK formula
        link_raw = col(row, "לינק מהיר") or ""
        file_path = get_file_path(row)

        database.insert_candidate({
            "file_hash": fake_hash,
            "file_path": file_path,
            "file_name": col(row, "שם קובץ") or "",
            "name": name,
            "email": email,
            "phone": phone,
            "current_role": col(row, "Current_Role"),
            "all_roles": [],
            "total_experience_years": years,
            "technologies": techs,
            "management_experience": mgmt,
            "management_years": None,
            "education": None,
            "location": None,
            "raw_summary": col(row, "Summary"),
        })
        imported += 1

    return {"imported": imported, "skipped": skipped, "message": f"יובאו {imported} מועמדים, דולגו {skipped} כפולים"}


class AnalyzeUrlRequest(BaseModel):
    url: str


@app.post("/api/analyze-job-url")
def analyze_job_url(req: AnalyzeUrlRequest):
    """
    Fetch a job listing URL and extract structured job requirements using Claude.
    """
    import urllib.request
    import html

    try:
        request = urllib.request.Request(
            req.url,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        )
        with urllib.request.urlopen(request, timeout=10) as resp:
            raw = resp.read().decode("utf-8", errors="ignore")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"לא ניתן לטעון את הכתובת: {e}")

    # Strip HTML tags
    clean = re.sub(r"<style[^>]*>.*?</style>", " ", raw, flags=re.DOTALL)
    clean = re.sub(r"<script[^>]*>.*?</script>", " ", clean, flags=re.DOTALL)
    clean = re.sub(r"<[^>]+>", " ", clean)
    clean = html.unescape(clean)
    clean = re.sub(r"\s+", " ", clean).strip()
    text = clean[:6000]

    prompt = f"""You are an expert tech recruiter. Read this job listing and extract the requirements.
Return ONLY a valid JSON object with these exact keys:
{{
  "role_title": "Job title",
  "min_experience_years": 5,
  "required_technologies": ["Python", "AWS"],
  "nice_to_have": ["Kubernetes"],
  "management_required": false,
  "location": "Tel Aviv or Remote",
  "domain": "FinTech",
  "org_type": "one of: לקוח קצה | אינטגרטור | ספק | פיננסי | ממשלתי | סטארטאפ | אנטרפרייז | or empty string",
  "salary_range": "salary range if mentioned, else empty string",
  "hybrid_mode": "one of: משרדי | היברידי | מרחוק | or empty string if unclear",
  "vendor_experience_required": false,
  "additional_notes": "Any other important requirements"
}}

Job listing text:
{text}
"""

    try:
        from anthropic import Anthropic
        from dotenv import load_dotenv
        load_dotenv()
        claude = Anthropic()
        msg = _claude_with_retry(lambda: claude.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=512,
            messages=[{"role": "user", "content": prompt}],
        ))
        response_text = msg.content[0].text.strip()
        if response_text.startswith("```"):
            response_text = "\n".join(response_text.split("\n")[1:-1])
        data = json.loads(response_text)
        return data
    except HTTPException:
        raise   # already user-friendly (from _claude_with_retry)
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="תגובת ה-AI לא הייתה בפורמט תקין. נסה שוב.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"שגיאה בניתוח: {e}")


@app.get("/api/candidates/{candidate_id}/history")
def get_candidate_history(candidate_id: int):
    return database.get_candidate_history(candidate_id)


@app.post("/api/candidates/{candidate_id}/open-cv")
def open_cv(candidate_id: int):
    """Open the candidate's CV file using the OS default application (Windows: os.startfile)."""
    import os as _os
    cand = database.get_candidate(candidate_id)
    if not cand:
        raise HTTPException(status_code=404, detail="מועמד לא נמצא")
    path = (cand.get("file_path") or "").strip()
    if not path:
        raise HTTPException(status_code=404, detail="אין קובץ קורות חיים מקושר למועמד זה")
    if not _os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"הקובץ לא נמצא בנתיב: {path}")
    try:
        _os.startfile(path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"לא ניתן לפתוח את הקובץ: {e}")
    return {"ok": True, "path": path}


# ─── Candidate notes ─────────────────────────────────────────────────────────

class AddNoteRequest(BaseModel):
    note: str
    author: str = ""


@app.get("/api/candidates/{candidate_id}/notes")
def get_notes(candidate_id: int):
    return database.get_candidate_notes(candidate_id)


@app.post("/api/candidates/{candidate_id}/notes")
def add_note(candidate_id: int, req: AddNoteRequest):
    if not req.note.strip():
        raise HTTPException(status_code=400, detail="הערה לא יכולה להיות ריקה")
    return database.add_candidate_note(candidate_id, req.note, req.author)


@app.delete("/api/candidates/{candidate_id}/notes/{note_id}")
def delete_note(candidate_id: int, note_id: int):
    database.delete_candidate_note(note_id)
    return {"ok": True}


# ─── Reminders ───────────────────────────────────────────────────────────────

class AddReminderRequest(BaseModel):
    title: str
    description: Optional[str] = ""
    due_at: str                          # ISO datetime string
    candidate_id: Optional[int] = None
    job_id: Optional[int] = None
    candidate_name: Optional[str] = ""
    job_title: Optional[str] = ""
    assigned_user_id: Optional[int] = None
    candidate_phone: Optional[str] = ""


@app.get("/api/reminders")
def get_reminders():
    return database.get_reminders(include_dismissed=True)


@app.get("/api/reminders/due")
def get_due_reminders():
    return database.get_due_reminders()


@app.post("/api/reminders")
def add_reminder(req: AddReminderRequest):
    if not req.title.strip():
        raise HTTPException(status_code=400, detail="כותרת לא יכולה להיות ריקה")
    if not req.due_at:
        raise HTTPException(status_code=400, detail="תאריך ושעה הם שדות חובה")
    return database.add_reminder(
        title=req.title, description=req.description or "",
        due_at=req.due_at, candidate_id=req.candidate_id, job_id=req.job_id,
        candidate_name=req.candidate_name or "", job_title=req.job_title or "",
        assigned_user_id=req.assigned_user_id,
        candidate_phone=req.candidate_phone or "",
    )


class UpdateReminderRequest(BaseModel):
    title: str
    description: Optional[str] = ""
    due_at: str
    candidate_name: Optional[str] = ""
    candidate_phone: Optional[str] = ""
    job_id: Optional[int] = None
    job_title: Optional[str] = ""
    assigned_user_id: Optional[int] = None


@app.put("/api/reminders/{reminder_id}")
def update_reminder(reminder_id: int, req: UpdateReminderRequest):
    if not req.title.strip():
        raise HTTPException(status_code=400, detail="כותרת לא יכולה להיות ריקה")
    result = database.update_reminder(
        reminder_id=reminder_id,
        title=req.title, description=req.description or "",
        due_at=req.due_at,
        candidate_name=req.candidate_name or "",
        candidate_phone=req.candidate_phone or "",
        job_id=req.job_id, job_title=req.job_title or "",
        assigned_user_id=req.assigned_user_id,
    )
    if not result:
        raise HTTPException(status_code=404, detail="תזכורת לא נמצאה או כבר הושלמה")
    return result


@app.patch("/api/reminders/{reminder_id}/dismiss")
def dismiss_reminder(reminder_id: int):
    database.dismiss_reminder(reminder_id)
    return {"ok": True}


@app.delete("/api/reminders/{reminder_id}")
def delete_reminder_endpoint(reminder_id: int):
    database.delete_reminder(reminder_id)
    return {"ok": True}


# ─── Users ───────────────────────────────────────────────────────────────────

class UserRequest(BaseModel):
    name: str
    role: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""


@app.get("/api/users")
def get_users():
    return database.get_users()


@app.post("/api/users")
def add_user(req: UserRequest):
    if not req.name.strip():
        raise HTTPException(status_code=400, detail="שם המשתמש לא יכול להיות ריק")
    return database.add_user(req.name, req.role or "", req.email or "", req.phone or "")


@app.put("/api/users/{user_id}")
def update_user(user_id: int, req: UserRequest):
    if not req.name.strip():
        raise HTTPException(status_code=400, detail="שם המשתמש לא יכול להיות ריק")
    database.update_user(user_id, req.name, req.role or "", req.email or "", req.phone or "")
    return {"ok": True}


@app.delete("/api/users/{user_id}")
def delete_user(user_id: int):
    database.delete_user(user_id)
    return {"ok": True}


# ─── Auth endpoints ───────────────────────────────────────────────────────────

class LoginRequest(BaseModel):
    email: str
    password: str


class SetPasswordRequest(BaseModel):
    password: str


@app.post("/api/auth/login")
def login(req: LoginRequest):
    user = database.get_user_by_email(req.email)
    if not user or not user.get("password_hash"):
        raise HTTPException(status_code=401, detail="שם משתמש או סיסמה שגויים")
    if not _pwd.verify(req.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="שם משתמש או סיסמה שגויים")
    token = _make_token(user["id"], user["email"])
    resp = JSONResponse({"ok": True, "user": {
        "id": user["id"], "name": user["name"],
        "role": user["role"], "email": user["email"],
    }})
    resp.set_cookie("aih_token", token, max_age=60 * 60 * 24 * _DAYS, httponly=True, samesite="lax")
    return resp


@app.post("/api/auth/logout")
def logout():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("aih_token")
    return resp


@app.get("/api/auth/me")
def get_me(request: Request):
    token = _get_token(request)
    payload = _decode_token(token) if token else None
    if not payload:
        raise HTTPException(status_code=401, detail="לא מורשה")
    user = database.get_user_by_id(payload["user_id"])
    if not user:
        raise HTTPException(status_code=401, detail="משתמש לא נמצא")
    return {"id": user["id"], "name": user["name"], "role": user["role"], "email": user["email"]}


@app.post("/api/users/{user_id}/set-password")
def set_password(user_id: int, req: SetPasswordRequest):
    if len(req.password) < 6:
        raise HTTPException(status_code=400, detail="הסיסמה חייבת להכיל לפחות 6 תווים")
    user = database.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="משתמש לא נמצא")
    database.set_user_password(user_id, _pwd.hash(req.password))
    return {"ok": True}



# ─── App settings ─────────────────────────────────────────────────────────────

class SmtpSettingsRequest(BaseModel):
    smtp_host: Optional[str] = "smtp.gmail.com"
    smtp_port: Optional[int] = 587
    smtp_user: Optional[str] = ""
    smtp_pass: Optional[str] = ""
    smtp_from_name: Optional[str] = "AI Head Hunter"


@app.get("/api/settings/smtp")
def get_smtp_settings():
    return {
        "smtp_host":      database.get_setting("smtp_host", "smtp.gmail.com"),
        "smtp_port":      int(database.get_setting("smtp_port", "587")),
        "smtp_user":      database.get_setting("smtp_user", ""),
        "smtp_pass":      database.get_setting("smtp_pass", ""),
        "smtp_from_name": database.get_setting("smtp_from_name", "AI Head Hunter"),
    }


@app.post("/api/settings/smtp")
def save_smtp_settings(req: SmtpSettingsRequest):
    database.set_setting("smtp_host",      req.smtp_host or "smtp.gmail.com")
    database.set_setting("smtp_port",      str(req.smtp_port or 587))
    database.set_setting("smtp_user",      req.smtp_user or "")
    database.set_setting("smtp_pass",      req.smtp_pass or "")
    database.set_setting("smtp_from_name", req.smtp_from_name or "AI Head Hunter")
    return {"ok": True}


@app.post("/api/settings/smtp/test")
def test_smtp(req: SmtpSettingsRequest):
    """Send a test email to the smtp_user address to verify settings."""
    if not req.smtp_user or not req.smtp_pass:
        raise HTTPException(status_code=400, detail="נא להזין כתובת מייל וסיסמה")
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = "🔔 AI Head Hunter — בדיקת חיבור מייל"
        msg["From"] = f"{req.smtp_from_name or 'AI Head Hunter'} <{req.smtp_user}>"
        msg["To"] = req.smtp_user
        msg.attach(MIMEText(
            "<div dir='rtl' style='font-family:sans-serif'><h2>✅ חיבור המייל עובד!</h2>"
            "<p>הגדרות ה-SMTP שלך נשמרו בהצלחה.</p></div>",
            "html", "utf-8",
        ))
        with smtplib.SMTP(req.smtp_host or "smtp.gmail.com", req.smtp_port or 587, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.login(req.smtp_user, req.smtp_pass)
            server.sendmail(req.smtp_user, [req.smtp_user], msg.as_string())
        return {"ok": True}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"שגיאה בחיבור: {exc}")


class UpdateCandidateNameRequest(BaseModel):
    name: str


@app.patch("/api/candidates/{candidate_id}/name")
def update_candidate_name(candidate_id: int, req: UpdateCandidateNameRequest):
    name = req.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="שם לא יכול להיות ריק")
    database.update_candidate_name(candidate_id, name)
    return {"ok": True, "name": name}


@app.patch("/api/candidates/{candidate_id}/extras")
def update_candidate_extras(candidate_id: int, req: UpdateCandidateExtrasRequest):
    """Update salary_expectation, availability, and/or call_notes for a candidate."""
    database.update_candidate_extras(
        candidate_id,
        salary_expectation=req.salary_expectation,
        availability=req.availability,
        call_notes=req.call_notes,
    )
    return {"ok": True}


class UpdateCandidateExtrasRequest(BaseModel):
    salary_expectation: Optional[str] = None
    availability: Optional[str] = None
    call_notes: Optional[str] = None


class CheckForJobRequest(BaseModel):
    job_id: int
    mode: Optional[str] = "standard"   # "standard" | "conservative" | "marketing"


class RejectCandidateRequest(BaseModel):
    candidate_id: int
    rejection_reason: str


class AcceptCandidateRequest(BaseModel):
    candidate_id: int
    acceptance_note: str = ""


@app.post("/api/jobs/{job_id}/reject-candidate")
def reject_candidate(job_id: int, req: RejectCandidateRequest):
    """Mark a candidate as rejected for this job with a reason."""
    database.reject_candidate_for_job(job_id, req.candidate_id, req.rejection_reason)
    return {"ok": True}


@app.post("/api/jobs/{job_id}/accept-candidate")
def accept_candidate(job_id: int, req: AcceptCandidateRequest):
    """Mark a candidate as accepted for this job with an optional note."""
    database.accept_candidate_for_job(job_id, req.candidate_id, req.acceptance_note)
    return {"ok": True}


@app.delete("/api/jobs/{job_id}/remove-candidate/{candidate_id}")
def remove_candidate_from_job(job_id: int, candidate_id: int):
    """Fully remove a candidate from a job's match history."""
    database.remove_candidate_from_job(job_id, candidate_id)
    return {"ok": True}


@app.post("/api/candidates/{candidate_id}/check-for-job")
def check_candidate_for_job(candidate_id: int, req: CheckForJobRequest):
    """ניתוח מעמיק של מועמד ספציפי מול משרה — כולל 3 סיכומי AI."""
    candidate = database.get_candidate_by_id(candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail="מועמד לא נמצא")

    job = database.get_job_by_id(req.job_id)
    if not job:
        raise HTTPException(status_code=404, detail="משרה לא נמצאה")

    try:
        result = job_matcher.check_candidate_for_job(candidate, job["requirements"], mode=req.mode or "standard")
    except RuntimeError as e:
        raise HTTPException(status_code=503, detail=str(e))
    if not result:
        raise HTTPException(status_code=500, detail="הניתוח נכשל — נסה שנית")

    # Extract AI summaries from result before saving match data
    summaries = {
        "submission_summary": result.pop("submission_summary", ""),
        "cv_review":          result.pop("cv_review", ""),
        "cv_improvements":    result.pop("cv_improvements", []),
        "short_summary":      result.pop("short_summary", ""),
        "missing_summary":    result.pop("missing_summary", []),
        "match_grades":       result.pop("match_grades", []),
    }

    database.save_match_result(
        candidate_id=candidate_id,
        job_id=req.job_id,
        job_title=job["title"],
        job_requirements=job["requirements"],
        result=result,
        summaries=summaries,
        match_type="single",
    )

    return {**result, "summaries": summaries, "job_title": job["title"]}


@app.post("/api/candidates/{candidate_id}/find-jobs")
def find_jobs_for_candidate(candidate_id: int):
    """מצא משרות תואמות למועמד — SSE streaming, ניתוח עמוק מול כל המשרות השמורות."""
    candidate = database.get_candidate_by_id(candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail="מועמד לא נמצא")

    jobs = database.get_all_jobs()

    def generate():
        if not jobs:
            yield f"data: {json.dumps({'type': 'done', 'total': 0, 'found': 0})}\n\n"
            return

        yield f"data: {json.dumps({'type': 'start', 'total_jobs': len(jobs)})}\n\n"

        found = 0
        for i, job in enumerate(jobs):
            try:
                # Merge requirements fields (role_title, technologies…) with job title
                job_reqs = {**job.get("requirements", {}), "title": job.get("title", "")}
                result = job_matcher.check_candidate_for_job(candidate, job_reqs)
                if result:
                    summaries = {
                        "submission_summary": result.pop("submission_summary", ""),
                        "cv_review":          result.pop("cv_review", ""),
                        "cv_improvements":    result.pop("cv_improvements", []),
                    }
                    database.save_match_result(
                        candidate_id=candidate_id,
                        job_id=job["id"],
                        job_title=job["title"],
                        job_requirements=job["requirements"],
                        result=result,
                        summaries=summaries,
                        match_type="find_jobs",
                    )
                    found += 1
                    event_data = {
                        "type": "result",
                        "job_id": job["id"],
                        "job_title": job["title"],
                        "processed": i + 1,
                        **result,
                        "summaries": summaries,
                    }
                    yield f"data: {json.dumps(event_data, ensure_ascii=False)}\n\n"
                else:
                    yield f"data: {json.dumps({'type': 'progress', 'processed': i + 1})}\n\n"
            except Exception as e:
                logger.error(f"find_jobs error for job {job['id']}: {e}", exc_info=True)
                yield f"data: {json.dumps({'type': 'progress', 'processed': i + 1, 'err': str(e)[:80]})}\n\n"

        yield f"data: {json.dumps({'type': 'done', 'total': len(jobs), 'found': found})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


@app.post("/api/candidates/{candidate_id}/interview-prep")
def interview_prep_for_candidate(candidate_id: int, req: CheckForJobRequest):
    """Generate interview preparation advice for a specific candidate + job."""
    candidate = database.get_candidate_by_id(candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail="מועמד לא נמצא")
    job = database.get_job_by_id(req.job_id)
    if not job:
        raise HTTPException(status_code=404, detail="משרה לא נמצאה")
    try:
        prep = job_matcher.generate_interview_prep(candidate, job["requirements"])
    except RuntimeError as e:
        raise HTTPException(status_code=503, detail=str(e))
    if not prep:
        raise HTTPException(status_code=500, detail="שגיאה ביצירת הכנה לראיון")
    return prep


class SubmissionBuilderRequest(BaseModel):
    job_text: str
    candidate_text: str
    candidate_name: Optional[str] = ""
    salary: Optional[str] = ""
    availability: Optional[str] = ""
    call_notes: Optional[str] = ""
    mode: Optional[str] = "standard"  # standard | conservative | marketing


@app.post("/api/submission-builder")
def submission_builder(req: SubmissionBuilderRequest):
    """ניתוח מלא ממשרה + קו"ח בפורמט חופשי — ללא צורך בייבוא למאגר."""
    if not req.job_text.strip():
        raise HTTPException(status_code=400, detail="תיאור המשרה ריק")
    if not req.candidate_text.strip():
        raise HTTPException(status_code=400, detail="פרטי המועמד ריקים")

    try:
        result = job_matcher.analyze_submission_raw(
            job_text=req.job_text,
            candidate_text=req.candidate_text,
            candidate_name=req.candidate_name or "",
            salary=req.salary or "",
            availability=req.availability or "",
            call_notes=req.call_notes or "",
            mode=req.mode or "standard",
        )
    except RuntimeError as e:
        raise HTTPException(status_code=503, detail=str(e))

    if not result:
        raise HTTPException(status_code=500, detail="הניתוח נכשל — נסה שנית")

    return result


@app.post("/api/enrich-summaries")
def enrich_summaries():
    """Translate existing English summaries to Hebrew for candidates missing raw_summary_he."""
    from anthropic import Anthropic
    from dotenv import load_dotenv
    load_dotenv()
    claude = Anthropic()

    missing = database.get_candidates_missing_he_summary()
    if not missing:
        return {"updated": 0, "message": "כל הסיכומים כבר בעברית"}

    def generate():
        updated = 0
        for c in missing:
            try:
                msg = claude.messages.create(
                    model="claude-haiku-4-5-20251001",
                    max_tokens=512,
                    messages=[{
                        "role": "user",
                        "content": f"Translate the following professional summary to fluent Hebrew. Return ONLY the Hebrew translation, nothing else.\n\n{c['raw_summary']}"
                    }]
                )
                he_text = msg.content[0].text.strip()
                database.update_summary_he(c["id"], he_text)
                updated += 1
                yield f"data: {json.dumps({'type': 'done', 'name': c['name'], 'updated': updated, 'total': len(missing)})}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'type': 'error', 'name': c['name'], 'error': str(e)})}\n\n"

        yield f"data: {json.dumps({'type': 'finished', 'updated': updated, 'total': len(missing)})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


@app.post("/api/reprocess-hebrew-candidates")
def reprocess_hebrew_candidates():
    """
    Re-extract structured data for candidates whose files still exist on disk
    and haven't been language-tagged yet. Detects Hebrew CVs and re-extracts
    with Hebrew role titles. SSE streaming with progress.
    """
    candidates = database.get_candidates_for_reprocess()

    def generate():
        if not candidates:
            yield f"data: {json.dumps({'type': 'finished', 'updated': 0, 'skipped': 0, 'total': 0, 'message': 'אין מועמדים לעדכון'})}\n\n"
            return

        yield f"data: {json.dumps({'type': 'start', 'total': len(candidates)})}\n\n"
        updated = 0
        skipped = 0

        for cand in candidates:
            cid    = cand["id"]
            name   = cand["name"]
            fpath  = cand.get("file_path") or ""
            fname  = cand.get("file_name") or ""

            if not fpath or not os.path.isfile(fpath):
                skipped += 1
                yield f"data: {json.dumps({'type': 'skip', 'name': name, 'reason': 'קובץ לא נמצא'})}\n\n"
                continue

            try:
                text = cv_scanner.extract_text(fpath)
            except Exception as e:
                skipped += 1
                yield f"data: {json.dumps({'type': 'skip', 'name': name, 'reason': f'שגיאת קריאה: {e}'})}\n\n"
                continue

            if not text or not cv_processor.detect_hebrew(text):
                # Not Hebrew — just mark as 'en' without re-extracting
                database.update_candidate_language(cid, "en")
                skipped += 1
                yield f"data: {json.dumps({'type': 'skip', 'name': name, 'reason': 'לא בעברית'})}\n\n"
                continue

            # Hebrew CV — re-extract with Hebrew prompt
            extracted = cv_processor.extract_candidate_data(text, fname)
            if not extracted:
                skipped += 1
                yield f"data: {json.dumps({'type': 'error', 'name': name, 'reason': 'חילוץ נכשל'})}\n\n"
                continue

            database.update_candidate_extracted(cid, extracted)
            updated += 1
            yield f"data: {json.dumps({'type': 'done', 'name': name, 'updated': updated, 'total': len(candidates)})}\n\n"

        yield f"data: {json.dumps({'type': 'finished', 'updated': updated, 'skipped': skipped, 'total': len(candidates)})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


# ─── Jobs ────────────────────────────────────────────────────────────────────

class SaveJobRequest(BaseModel):
    job_requirements: JobRequirements


@app.get("/api/jobs")
def list_jobs():
    return database.get_all_jobs()


@app.post("/api/jobs")
def create_job(req: SaveJobRequest):
    job_dict = req.job_requirements.model_dump()
    # Generate reference number: DDMMYYYY/N where N = today's job count + 1
    today = datetime.now()
    n = database.count_jobs_today() + 1
    reference_number = today.strftime("%d%m%Y") + f"/{n}"
    job_id = database.save_job(req.job_requirements.role_title, job_dict, reference_number)
    return {"id": job_id, "title": req.job_requirements.role_title, "reference_number": reference_number}


@app.put("/api/jobs/{job_id}")
def update_job(job_id: int, req: SaveJobRequest):
    job_dict = req.job_requirements.model_dump()
    database.update_job(job_id, req.job_requirements.role_title, job_dict)
    return {"ok": True}


@app.delete("/api/jobs/{job_id}")
def remove_job(job_id: int):
    database.delete_job(job_id)
    return {"ok": True}


@app.delete("/api/jobs/{job_id}/history")
def clear_job_history(job_id: int):
    """מחק את כל היסטוריית ההתאמות של משרה ספציפית."""
    database.clear_job_history(job_id)
    return {"ok": True}


# ── Civi import endpoints ──────────────────────────────────────────────────────

@app.get("/api/jobs/pending")
def list_pending_jobs():
    """Return jobs imported from Civi that are awaiting user approval."""
    return database.get_pending_jobs()


@app.post("/api/jobs/import-civi")
def import_civi_jobs():
    """Scrape Civi job listing and save new jobs as 'pending'."""
    try:
        scraped = civi_scraper.scrape_civi_jobs()
    except Exception as e:
        logging.error(f"Civi scrape failed: {e}")
        raise HTTPException(status_code=502, detail=f"לא ניתן לגשת לאתר Civi: {e}")

    added, skipped = 0, 0
    for job in scraped:
        new_id = database.save_pending_job(
            title=job["title"],
            civi_id=job["civi_id"],
            civi_url=job["url"],
        )
        if new_id:
            added += 1
        else:
            skipped += 1

    return {"added": added, "skipped": skipped, "total_scraped": len(scraped)}


def _fetch_and_analyze_job_url(url: str, fallback_title: str) -> dict:
    """Fetch a job listing URL, strip HTML, run Claude, and return requirements dict."""
    import urllib.request as _urlib
    import html as _html

    try:
        req = _urlib.Request(
            url,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        )
        with _urlib.urlopen(req, timeout=10) as resp:
            raw = resp.read().decode("utf-8", errors="ignore")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"לא ניתן לטעון את דף המשרה: {e}")

    clean = re.sub(r"<style[^>]*>.*?</style>", " ", raw, flags=re.DOTALL)
    clean = re.sub(r"<script[^>]*>.*?</script>", " ", clean, flags=re.DOTALL)
    clean = re.sub(r"<[^>]+>", " ", clean)
    clean = _html.unescape(clean)
    clean = re.sub(r"\s+", " ", clean).strip()[:6000]

    prompt = f"""You are an expert tech recruiter. Read this job listing and extract the requirements.
Return ONLY a valid JSON object with these exact keys:
{{
  "role_title": "Job title",
  "min_experience_years": 5,
  "required_technologies": ["Python", "AWS"],
  "nice_to_have": ["Kubernetes"],
  "management_required": false,
  "location": "Tel Aviv or Remote",
  "domain": "FinTech",
  "org_type": "one of: לקוח קצה | אינטגרטור | ספק | פיננסי | ממשלתי | סטארטאפ | אנטרפרייז | or empty string",
  "salary_range": "salary range if mentioned, else empty string",
  "hybrid_mode": "one of: משרדי | היברידי | מרחוק | or empty string if unclear",
  "vendor_experience_required": false,
  "additional_notes": "Any other important requirements"
}}

Job listing text:
{clean}
"""

    try:
        from anthropic import Anthropic
        from dotenv import load_dotenv
        load_dotenv()
        claude = Anthropic()
        msg = _claude_with_retry(lambda: claude.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=512,
            messages=[{"role": "user", "content": prompt}],
        ))
        response_text = msg.content[0].text.strip()
        if response_text.startswith("```"):
            response_text = "\n".join(response_text.split("\n")[1:-1])
        return json.loads(response_text)
    except HTTPException:
        raise
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="תגובת ה-AI לא הייתה בפורמט תקין. נסה שוב.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"שגיאה בניתוח המשרה: {e}")


@app.post("/api/jobs/{job_id}/analyze")
def analyze_pending_job(job_id: int):
    """
    Analyze a pending Civi job with Claude but keep it as pending.
    Returns the extracted requirements so the user can review/edit before approving.
    """
    job = database.get_job_by_id(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="משרה לא נמצאה")

    civi_url = job.get("civi_url")
    if not civi_url:
        raise HTTPException(status_code=400, detail="אין קישור Civi למשרה זו")

    requirements = _fetch_and_analyze_job_url(civi_url, job["title"])

    # Save requirements to the pending job (keeps status=pending)
    title = requirements.get("role_title") or job["title"]
    database.update_job(job_id, title, requirements)

    return {"ok": True, "title": title, "requirements": requirements}


class ApproveJobRequest(BaseModel):
    requirements: Optional[dict] = None  # if provided, skip Claude analysis


@app.post("/api/jobs/{job_id}/approve")
def approve_job(job_id: int, req: Optional[ApproveJobRequest] = None):
    """
    Approve a pending Civi job and set it to active.
    If req.requirements is provided (user already analyzed+edited), use those directly.
    Otherwise, run Claude analysis first.
    """
    job = database.get_job_by_id(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="משרה לא נמצאה")

    # Use pre-analyzed requirements if provided; otherwise analyze now
    if req and req.requirements:
        requirements = req.requirements
    else:
        civi_url = job.get("civi_url")
        if not civi_url:
            raise HTTPException(status_code=400, detail="אין קישור Civi למשרה זו")
        requirements = _fetch_and_analyze_job_url(civi_url, job["title"])

    title = requirements.get("role_title") or job["title"]
    database.approve_pending_job(job_id, title, requirements)
    return {"ok": True, "title": title, "requirements": requirements}


@app.post("/api/jobs/{job_id}/freeze")
def freeze_job(job_id: int):
    """Set job status to frozen (hidden from active list but kept in DB)."""
    database.update_job_status(job_id, "frozen")
    return {"ok": True}


@app.post("/api/jobs/{job_id}/unfreeze")
def unfreeze_job(job_id: int):
    """Restore a frozen job to active."""
    database.update_job_status(job_id, "active")
    return {"ok": True}


@app.post("/api/jobs/{job_id}/close")
def close_pending_job(job_id: int):
    """Permanently remove a pending Civi job (before it's approved)."""
    database.delete_job(job_id)
    return {"ok": True}


@app.delete("/api/history")
def clear_all_history():
    """מחק את כל היסטוריית ההתאמות של כל המשרות."""
    database.clear_all_history()
    return {"ok": True}


@app.get("/api/jobs/{job_id}/matches")
def get_job_matches(job_id: int):
    return database.get_job_matches(job_id)


@app.get("/api/jobs/{job_id}/runs")
def get_match_runs(job_id: int):
    return database.get_match_runs(job_id)


@app.post("/api/jobs/{job_id}/match")
async def match_for_job(job_id: int, req: MatchRequest):
    """התאם מועמדים למשרה שמורה — SSE streaming עם progress.
    רץ ב-background thread כדי שהשרת לא ייחסם, עם keepalive כל 15 שניות למניעת disconnect."""
    candidates = database.get_all_candidates()
    job_dict = req.job_requirements.model_dump()
    job_title = req.job_requirements.role_title

    # Always exclude rejected candidates (regardless of forceIds)
    rejected_ids = database.get_rejected_candidate_ids(job_id)
    if rejected_ids:
        candidates = [c for c in candidates if c["id"] not in rejected_ids]

    # Pre-load accepted candidates — they'll be forced into results at top
    accepted_candidates = database.get_accepted_candidates_full(job_id)
    accepted_ids = {a["id"] for a in accepted_candidates}

    if req.candidate_ids:
        id_set = set(req.candidate_ids) - accepted_ids
        candidates = [c for c in candidates if c["id"] in id_set]
    else:
        already_matched = database.get_matched_candidate_ids(job_id)
        skip_ids = (already_matched | accepted_ids)
        if skip_ids:
            candidates = [c for c in candidates if c["id"] not in skip_ids]

    if not candidates and not accepted_candidates:
        async def _empty():
            yield f"data: {json.dumps({'type': 'results', 'results': [], 'total_candidates': 0, 'candidates_checked': 0, 'matches_found': 0, 'message': 'אין מועמדים במאגר'})}\n\n"
        return StreamingResponse(_empty(), media_type="text/event-stream")

    rejections = database.get_rejections_for_job(job_id)
    effective_top_n   = req.top_n       if req.top_n       is not None else (job_dict.get("top_n")       or 10)
    effective_min_pct = req.min_match_pct if req.min_match_pct is not None else (job_dict.get("min_match_pct") or 60)

    loop = asyncio.get_event_loop()
    q: asyncio.Queue = asyncio.Queue()

    def _sync_worker():
        """Runs the heavy AI matching in a background thread — never blocks the event loop."""
        try:
            if not candidates:
                final_results = accepted_candidates
                for r in final_results:
                    cid = r.get("id")
                    if cid:
                        database.save_match_result(cid, job_id, job_title, job_dict, r)
                payload = json.dumps({
                    'type': 'results', 'results': final_results,
                    'total_candidates': len(accepted_candidates),
                    'candidates_checked': 0, 'matches_found': len(final_results),
                }, ensure_ascii=False)
                loop.call_soon_threadsafe(q.put_nowait, payload)
                return

            for event in job_matcher.match_candidates_streaming(
                job=job_dict,
                candidates=candidates,
                top_n=effective_top_n,
                min_match_pct=effective_min_pct,
                rejections=rejections,
                skip_pre_filter=bool(req.candidate_ids),
            ):
                if event["type"] == "results":
                    final_results = accepted_candidates + event["results"]
                    for r in final_results:
                        cid = r.get("id")
                        if cid:
                            database.save_match_result(cid, job_id, job_title, job_dict, r)
                    database.log_match_run(
                        job_id=job_id,
                        candidates_total=event["total_candidates"],
                        candidates_checked=event["candidates_checked"],
                        matches_found=len(final_results),
                        min_match_pct=req.min_match_pct,
                    )
                    payload = json.dumps({**event, 'results': final_results, 'matches_found': len(final_results)}, ensure_ascii=False)
                else:
                    payload = json.dumps(event, ensure_ascii=False)
                loop.call_soon_threadsafe(q.put_nowait, payload)
        except Exception as exc:
            logger.error("match_for_job background worker error: %s", exc, exc_info=True)
            loop.call_soon_threadsafe(q.put_nowait, json.dumps({"type": "error", "message": str(exc)[:300]}))
        finally:
            loop.call_soon_threadsafe(q.put_nowait, None)  # sentinel — signals end of stream

    threading.Thread(target=_sync_worker, daemon=True).start()

    async def _async_gen():
        """Async generator: reads from queue, sends keepalive every 15s to prevent SSE disconnect."""
        while True:
            try:
                item = await asyncio.wait_for(q.get(), timeout=15.0)
                if item is None:
                    break
                yield f"data: {item}\n\n"
            except asyncio.TimeoutError:
                yield ": keepalive\n\n"  # SSE comment — keeps connection alive, ignored by frontend

    return StreamingResponse(_async_gen(), media_type="text/event-stream")


QUICK_CHUNK = 500  # candidates per SSE batch for quick match

@app.post("/api/jobs/{job_id}/quick-match")
async def quick_match_for_job(job_id: int, req: MatchRequest):
    """התאמה מהירה — SSE streaming, ציון מיידי מבוסס מילות מפתח ללא AI.
    רץ ב-background thread כדי שהשרת לא ייחסם."""
    candidates = database.get_all_candidates()
    job_dict = req.job_requirements.model_dump()
    job_title = req.job_requirements.role_title
    effective_min_pct = req.min_match_pct if req.min_match_pct is not None else (job_dict.get("min_match_pct") or 60)
    effective_top_n   = req.top_n       if req.top_n       is not None else (job_dict.get("top_n")       or 20)
    min_exp = float(job_dict.get("min_experience_years") or 0)

    if req.candidate_ids:
        id_set = set(req.candidate_ids)
        candidates = [c for c in candidates if c["id"] in id_set]

    rejected_ids = database.get_rejected_candidate_ids(job_id)
    if rejected_ids:
        candidates = [c for c in candidates if c["id"] not in rejected_ids]

    accepted_candidates = database.get_accepted_candidates_full(job_id)
    accepted_ids = {a["id"] for a in accepted_candidates}

    exp_filter = min_exp == 0 or bool(req.candidate_ids)
    to_check = [
        c for c in candidates
        if c["id"] not in accepted_ids
        and (exp_filter or (c.get("total_experience_years") or 0) >= max(0, min_exp - 7))
    ]
    total_batches = max(1, (len(to_check) + QUICK_CHUNK - 1) // QUICK_CHUNK)

    loop = asyncio.get_event_loop()
    q: asyncio.Queue = asyncio.Queue()

    def _sync_worker():
        try:
            loop.call_soon_threadsafe(q.put_nowait, json.dumps({
                'type': 'start', 'total_candidates': len(candidates),
                'candidates_checked': len(to_check), 'total_batches': total_batches,
            }))

            if accepted_candidates:
                loop.call_soon_threadsafe(q.put_nowait, json.dumps({
                    'type': 'batch', 'batch_num': 0, 'total_batches': total_batches,
                    'processed_so_far': 0, 'candidates_checked': len(to_check),
                    'batch_results': accepted_candidates,
                }, ensure_ascii=False))

            all_results = []
            for batch_num, i in enumerate(range(0, len(to_check), QUICK_CHUNK), 1):
                chunk = to_check[i: i + QUICK_CHUNK]
                batch_results = []
                for c in chunk:
                    scored = quick_matcher._score(c, job_dict)
                    if scored["match_percentage"] >= effective_min_pct:
                        batch_results.append({**c, **scored})
                all_results.extend(batch_results)
                loop.call_soon_threadsafe(q.put_nowait, json.dumps({
                    'type': 'batch', 'batch_num': batch_num, 'total_batches': total_batches,
                    'processed_so_far': min(i + QUICK_CHUNK, len(to_check)),
                    'candidates_checked': len(to_check), 'batch_results': batch_results,
                }, ensure_ascii=False))

            all_results.sort(key=lambda x: x.get("match_percentage", 0), reverse=True)
            regular_results = all_results[:effective_top_n]
            top_results = accepted_candidates + regular_results

            for r in top_results:
                cid = r.get("id")
                if cid:
                    database.save_match_result(cid, job_id, job_title, job_dict, r, match_type="quick")

            logger.info(
                f"Quick match job={job_id} ({job_title!r}): total={len(candidates)}, "
                f"min_pct={effective_min_pct}, above_threshold={len(all_results)}, top_n={effective_top_n}, returned={len(top_results)}"
            )
            loop.call_soon_threadsafe(q.put_nowait, json.dumps({
                'type': 'results', 'total_candidates': len(candidates),
                'candidates_checked': len(to_check), 'above_threshold': len(all_results),
                'matches_found': len(top_results), 'top_n': effective_top_n,
            }))
        except Exception as exc:
            logger.error("quick_match_for_job background worker error: %s", exc, exc_info=True)
            loop.call_soon_threadsafe(q.put_nowait, json.dumps({"type": "error", "message": str(exc)[:300]}))
        finally:
            loop.call_soon_threadsafe(q.put_nowait, None)

    threading.Thread(target=_sync_worker, daemon=True).start()

    async def _async_gen():
        while True:
            try:
                item = await asyncio.wait_for(q.get(), timeout=15.0)
                if item is None:
                    break
                yield f"data: {item}\n\n"
            except asyncio.TimeoutError:
                yield ": keepalive\n\n"

    return StreamingResponse(_async_gen(), media_type="text/event-stream")


@app.post("/api/match")
def match_candidates(req: MatchRequest):
    """התאמה ללא שמירת משרה."""
    candidates = database.get_all_candidates()
    if not candidates:
        return {"results": [], "message": "אין מועמדים במאגר"}

    job_dict = req.job_requirements.model_dump()
    results = job_matcher.match_candidates(
        job=job_dict,
        candidates=candidates,
        top_n=req.top_n,
        min_match_pct=req.min_match_pct,
    )

    return {
        "results": results,
        "total_candidates_checked": len(candidates),
        "matches_found": len(results),
    }


# ─── Single CV add ────────────────────────────────────────────────────────────

def _process_cv_bytes(content: bytes, fname: str) -> dict:
    """Shared logic: extract text, call Claude, insert to DB. Returns candidate dict."""
    import tempfile

    ext = os.path.splitext(fname)[1].lower()
    if ext not in {".pdf", ".docx", ".doc"}:
        raise HTTPException(status_code=400, detail="סוג קובץ לא נתמך — רק PDF או DOCX")

    md5 = hashlib.md5(content).hexdigest()
    if database.hash_exists(md5):
        raise HTTPException(status_code=409, detail="קורות חיים אלו כבר קיימים במאגר")

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    try:
        text = cv_scanner.extract_text(tmp_path)
    finally:
        os.unlink(tmp_path)

    if not text.strip():
        raise HTTPException(status_code=422, detail="לא ניתן לחלץ טקסט — ייתכן PDF סרוק (תמונה בלבד)")

    extracted = cv_processor.extract_candidate_data(text, fname)
    if not extracted:
        raise HTTPException(status_code=422, detail="שגיאה בחילוץ נתוני המועמד")

    name  = extracted.get("name") or ""
    email = extracted.get("email") or ""
    existing = database.candidate_exists_by_identity(name, email)
    if existing:
        raise HTTPException(status_code=409, detail=f"מועמד זה כבר קיים במאגר בשם: {existing['name']} ({existing['file_name']})")

    extracted["file_hash"] = md5
    extracted["file_path"] = ""
    extracted["file_name"] = fname
    database.insert_candidate(extracted)
    return extracted


@app.post("/api/upload-cv")
async def upload_cv(file: UploadFile = File(...)):
    """הוסף מועמד יחיד מקובץ מועלה (PDF/DOCX)."""
    content = await file.read()
    candidate = _process_cv_bytes(content, file.filename or "cv.pdf")
    return {"success": True, "candidate": candidate}


class AddCvPathRequest(BaseModel):
    path_or_url: str


@app.post("/api/add-cv-path")
def add_cv_from_path(req: AddCvPathRequest):
    """הוסף מועמד יחיד לפי נתיב מקומי (G:\\...) או קישור Google Drive ציבורי."""
    src = req.path_or_url.strip()

    # Google Drive sharing URL → direct download
    if "drive.google.com" in src:
        m = re.search(r'/file/d/([^/?]+)', src)
        if not m:
            raise HTTPException(status_code=400, detail="לא ניתן לחלץ ID מהקישור. ודא שהקישור הוא של קובץ ספציפי.")
        file_id = m.group(1)
        download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
        import urllib.request as urlreq
        try:
            req2 = urlreq.Request(download_url, headers={"User-Agent": "Mozilla/5.0"})
            with urlreq.urlopen(req2, timeout=30) as resp:
                content = resp.read()
                ct = resp.headers.get("Content-Type", "")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"שגיאה בהורדת הקובץ: {e}. ודא שהקובץ שותף ציבורית.")
        # Guess extension from content-type
        ext = ".pdf" if "pdf" in ct else ".docx"
        fname = f"gdrive_{file_id}{ext}"
        candidate = _process_cv_bytes(content, fname)
        return {"success": True, "candidate": candidate}

    # Local path
    if not os.path.isfile(src):
        raise HTTPException(status_code=404, detail=f"קובץ לא נמצא: {src}")
    with open(src, "rb") as f:
        content = f.read()
    fname = os.path.basename(src)
    candidate = _process_cv_bytes(content, fname)
    candidate["file_path"] = src  # keep real path for local files
    database.get_conn()  # update file_path
    # Update the file_path we just inserted
    conn = database.get_conn()
    conn.execute("UPDATE candidates SET file_path = ? WHERE file_hash = ?",
                 (src, candidate["file_hash"]))
    conn.commit()
    conn.close()
    return {"success": True, "candidate": candidate}


# ─── AI Candidate Search ─────────────────────────────────────────────────────

_AI_SEARCH_PROMPT = """\
You are a search query builder for a Hebrew/English CV database.
Convert the recruiter's free-text query into a compound search expression.

RULES:
- Use '+' to separate AND conditions (ALL must match)
- Use '|' within an AND group to list OR alternatives (ANY one can match)
- Include BOTH Hebrew AND English variants for the same concept in the same OR group
- For roles: include both Hebrew and English names as alternatives
- For technologies: include alternate names (e.g. "flow|power automate|microsoft flow")
- Keep each OR group focused on ONE concept
- Maximum 6 AND groups, up to 4 OR alternatives per group
- "הכי חשוב"/"חובה"/"must" → make that concept a mandatory AND group
- "עדיף"/"nice to have" → still include as AND group (user can remove it)

TRANSLATION GUIDE:
- מנתח/ת מערכות → systems analyst
- מנהל/ת פרויקטים → project manager
- מפתח/ת → developer
- מהנדס/ת תוכנה → software engineer
- בודק/ת / QA → qa engineer
- Flow (Microsoft context) → flow|power automate|microsoft flow
- Power Platform → power platform|power apps|power automate

EXAMPLES:
Query: "אני צריך מנתחי מערכות שמכירים Power Platform. הכי חשוב שיכירו Flow."
Output: מנתח מערכות|systems analyst + power platform|power apps + flow|power automate|microsoft flow

Query: "Project managers or developers who know Python and AWS"
Output: מנהל פרויקטים|project manager|מפתח|developer + python + aws|amazon web services

Query: "מפתחים בכירים עם ניסיון ב-Salesforce, עדיף גם CRM"
Output: מפתח|developer|software engineer + salesforce + crm

IMPORTANT: Return ONLY the compound search expression on a single line. No explanation, no markdown, no quotes."""


class AISearchRequest(BaseModel):
    query: str


@app.post("/api/candidates/ai-search")
def ai_search_candidates(req: AISearchRequest):
    """Parse free-text recruiter query into compound AND/OR search tokens for client-side filtering."""
    if not req.query.strip():
        raise HTTPException(status_code=400, detail="שאילתת החיפוש ריקה")

    import anthropic as _anth
    _key = os.environ.get("ANTHROPIC_API_KEY", "").strip().encode("ascii", errors="ignore").decode("ascii")
    _cli = _anth.Anthropic(api_key=_key)

    prompt_text = _AI_SEARCH_PROMPT + f"\n\nQuery: {req.query.strip()}\nOutput:"

    def _call_claude():
        msg = _cli.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=200,
            messages=[{"role": "user", "content": prompt_text}],
        )
        return msg.content[0].text.strip()

    try:
        raw = _claude_with_retry(_call_claude)

        # Take the last non-empty line (model may add a brief intro sentence)
        lines = [l.strip() for l in raw.splitlines() if l.strip()]
        # Prefer a line containing '+' or '|', otherwise use the last line
        search_query = next(
            (l for l in reversed(lines) if '+' in l or '|' in l),
            lines[-1] if lines else ""
        )

        # Parse into token groups: outer list = AND, inner list = OR alternatives
        tokens: list[list[str]] = []
        for group in search_query.split('+'):
            alternatives = [a.strip() for a in group.split('|') if a.strip()]
            if alternatives:
                tokens.append(alternatives)

        return {"search_query": search_query, "tokens": tokens}

    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"ai_search_candidates error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"שגיאה בחיפוש: {str(e)[:120]}")


# ─── Serve built frontend (for Tailscale / production access) ────────────────

# Locate the built React frontend.
# • Frozen (PyInstaller): launcher embeds dist/ alongside the .py files in sys._MEIPASS
# • Dev: dist is at ../../frontend/dist relative to backend/main.py
def _find_dist() -> str:
    if getattr(sys, 'frozen', False):
        return os.path.join(sys._MEIPASS, 'dist')
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'frontend', 'dist')

_DIST = _find_dist()

if os.path.isdir(_DIST):
    app.mount("/assets", StaticFiles(directory=os.path.join(_DIST, "assets")), name="assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    def spa_fallback(full_path: str):
        """Serve index.html for all non-API routes (SPA client-side routing)."""
        index_path = os.path.join(_DIST, "index.html")
        with open(index_path, "rb") as f:
            content = f.read()
        return Response(
            content=content,
            media_type="text/html",
            headers={
                "Cache-Control": "no-store, no-cache, must-revalidate",
                "Pragma": "no-cache",
            },
        )
